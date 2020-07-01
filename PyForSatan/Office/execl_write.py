from openpyxl import Workbook


class execl_write:

    __data__ = {}

    resource = None

    def __init__(self):
        self.resource = Workbook()

    def load_data_to_file(self,sheet_names=None,data=None,data_type_str=None):
        """
        加载数据
        :param sheet_names: 当sheet_names 为 True 时 激活默认页
        :param data: 数据
        :param data_type_str: 数据类型 请参见__load_data_to_sheet方法
        :return:
        """

        sheet_page = self.__choose_sheets(sheet_names)

        sheet_pages_source = self.__load_data_to_sheet(sheet=sheet_page,data=data,data_Type=data_type_str)


        return self

    def __choose_sheets(self,sheet_names=None):
        """
        创建sheet页
        :param sheet_names: 当sheet_names 为 True 时 激活默认页
        :return:
        """
        if sheet_names is None:
            sheet = self.resource.create_sheet()
        elif sheet_names  == True:
            sheet = self.resource.active
        else:
            sheet = self.resource.create_sheet(sheet_names)


        return sheet

    def __load_data_to_sheet(self,sheet=None,data=None,data_Type=None):
        """
        把数据载入sheet页中
        :param sheet:
        :param data:
        :param data_Type:
        :return:
        """
        if data_Type is None:

            keys,values = self.__load_data_json(data)

        elif data_Type == "data_json":

            keys,values = self.__load_data_json(data)

        elif data_Type == "data_excel":

            keys,values = self.__load_data_yuanshi(data)

        else:
            raise RuntimeError("暂无此类方法")

        sheet.append(keys)

        for i in values:
            sheet.append(i)

        return sheet



    def __load_data_json(self,data):
        """
        读取第一行json的key做为标头 ，遍历循环 当没有该值是填充一个空的字符串
        :param data:
        :return: keys 表头   values  一行一行的值
        """
        if len(data) > 0 :
            keys = data[0].keys()
            data_temp = []
            for i in data:
                data_temp_hang = []
                for j in keys:
                    if j in i.keys():
                        data_temp_hang.append(i[j])
                    else:
                        data_temp_hang.append("")
                data_temp.append(data_temp_hang)
            return list(keys), data_temp
        else:
            return [], []

    def __load_data_yuanshi(self,data):

        if len(data) > 0 :
            keys = []
            for i in data[0]:
                 keys.extend(i.keys())

            values = []
            for i in data:
                values_hang = []
                for k in i:
                    values_hang.extend(list(k.values()))
                values.append(values_hang)

            return keys,values
        else:
            return [],[]


    def save_file(self,path=None):
        """
        保存。。生成文件
        :param path:
        :return: True or False
        """
        if path is not None:
            return self.resource.save(path)
        else:
            raise IOError("创建的execl路径不得为None..")


    def close(self):
        """
        关闭资源
        :return:
        """
        if self.resource is not None:
            self.resource.close()

    def __enter__(self):
        """
        兼容with方法
        :return:
        """
        pass
    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        兼容with方法
        :param exc_type:
        :param exc_val:
        :param exc_tb:
        :return:
        """
        if exc_tb is None:
            self.close()